# Ke Yan, Imaging Biomarkers and Computer-Aided Diagnosis Laboratory,
# National Institutes of Health Clinical Center, July 2019
"""Procedure in the demo mode"""
import os
import numpy as np
from time import time
import torch
import nibabel as nib
from tqdm import tqdm
import cv2
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd

from maskrcnn.config import cfg
from maskrcnn.data.datasets.load_ct_img import load_prep_img
from maskrcnn.structures.image_list import to_image_list
from maskrcnn.data.datasets.evaluation.DeepLesion.post_process import post_process_results
from maskrcnn.data.datasets.load_ct_img import windowing, windowing_rev
from maskrcnn.utils.draw import draw_results


def exec_model(model):
    """test model on user-provided data, instead of the preset DeepLesion dataset"""
    import_tag_data()
    model.eval()
    device = torch.device(cfg.MODEL.DEVICE)

    #while True:
        #info = "Please input the path of a nifti CT volume >> "
        #while True:
            #path = input(info)
# ------- Zhoubing 100 datasets -------
#    for num in range(12):
#        if num + 1  < 10:       
#            img_num = 'img000' + str(num + 1)
#            #data_dir = '/nfs/masi/leeh43/MULAN_universal_lesion_analysis/results'
#            #img_dir = '_nfs_masi_leeh43_zhoubing100_img_' + img_num + '.nii.gz/'
#            #result = os.path.join(data_dir, img_dir + 'results.txt' )
#            main_dir = '/nfs/masi/leeh43/zhoubing100/img/'
#            img_dir = os.path.join(main_dir, img_num + '.nii.gz')
#            
#        if num + 1 >= 10 and num + 1 < 100:       
#            img_num = 'img00' + str(num + 1)
#            main_dir = '/nfs/masi/leeh43/zhoubing100/img/'
#            img_dir = os.path.join(main_dir, img_num + '.nii.gz')
#            
#        if num + 1 == 100:       
#            img_num = 'img0' + str(num + 1)
#            main_dir = '/nfs/masi/leeh43/zhoubing100/img/'
#            img_dir = os.path.join(main_dir, img_num + '.nii.gz')
#            if not os.path.exists(img_dir):
#                print('file does not exist!')
#                continue
#        #try:
            
# ------- ImageVU B Datasets -------
    data_dir = os.path.join('/nfs/masi/tangy5/ImageVU_B_bpr_pipeline/INPUTS/cropped/images')
    count = 0
    for item in os.listdir(data_dir):
        img_dir = os.path.join(data_dir, item)
        
        print('reading image ...')
        nifti_data = nib.load(img_dir)
        count = count + 1
        print('Number of Datasets: %d' % count)
        print('Load Image: %s' % img_dir)
            #break
        #except:
            #print('load nifti file error!')

        while True:
            win_sel = '1' #input('Window to show, 1:soft tissue, 2:lung, 3: bone >> ')
            if win_sel not in ['1', '2', '3']:
                continue
            win_show = [[-175, 275], [-1500, 500], [-500, 1300]]
            win_show = win_show[int(win_sel)-1]
            break

        vol, spacing, slice_intv = load_preprocess_nifti(nifti_data)
            
        slice_num_per_run = max(1, int(float(cfg.TEST.TEST_SLICE_INTV_MM)/slice_intv+.5))
        num_total_slice = vol.shape[2]

        total_time = 0
        imageVU_dir = 'ImageVU_B_result'
        output_dir = os.path.join(cfg.RESULTS_DIR,imageVU_dir,img_dir.replace(os.sep, '_'))
        if not os.path.exists(output_dir):
            os.mkdir(output_dir)
        
        
        slices_to_process = range(int(slice_num_per_run/2), num_total_slice, slice_num_per_run)
        msgs_all = []
        print('predicting ...')
        for slice_idx in tqdm(slices_to_process):
            log_file_s = os.path.join(output_dir, 'slice_' + str(slice_idx) + '_resize_shape.csv') 
            log_file_c = os.path.join(output_dir, 'slice_' + str(slice_idx) + '_contour_location.csv')    
            log_file_r = os.path.join(output_dir, 'slice_' + str(slice_idx) + '_recist_location.csv')
            log_file_mask = os.path.join(output_dir, 'slice_' + str(slice_idx) + '_mask_c.csv')
            mask_list = []
            ims, im_np, im_scale, crop, mask_list = get_ims(slice_idx, vol, spacing, slice_intv, mask_list)
            im_list = to_image_list(ims, cfg.DATALOADER.SIZE_DIVISIBILITY).to(device)
            start_time = time()
            with torch.no_grad():
                result = model(im_list)
            result = [o.to("cpu") for o in result]
            
            df_resize = pd.DataFrame()
            df_contours = pd.DataFrame()
            df_recists = pd.DataFrame()
            df_mask = pd.DataFrame()
            shape_0, shape_1 = [], []
            cour_list1, cour_list2 = [], []
            recist_list1, recist_list2 = [], []
            info = {'spacing': spacing, 'im_scale': im_scale}
            post_process_results(result[0], info)
            total_time += time() - start_time
            output_fn = os.path.join(output_dir, '%d.png'%(slice_idx+1))
            real_slice_num = slice_idx + 1
            #contour_list.append('Slice_'+str(real_slice_num))
            #recist_list.append(('Slice_'+str(real_slice_num)))
            shape_0.append(im_np.shape[0])
            shape_1.append(im_np.shape[1])
            overlay, msgs = gen_output(im_np, result[0], info, win_show, cour_list1, cour_list2, recist_list1, recist_list2)
            df_resize['Shape_0'] = shape_0
            df_resize['Shape_1'] = shape_1
            df_contours['list1'] = cour_list1
            df_contours['list2'] = cour_list2
            df_recists['list1'] = recist_list1
            df_mask['c'] = mask_list
            df_resize.to_csv(log_file_s, index=False)
            df_contours.to_csv(log_file_c, index = False)
            df_recists.to_csv(log_file_r, index = False)
            df_mask.to_csv(log_file_mask, index = False)
            cv2.imwrite(output_fn, overlay)
            msgs_all.append('slice %d\r\n' % (slice_idx+1))
            for msg in msgs:
                msgs_all.append(msg+'\r\n')
            msgs_all.append('\r\n')
        
        #np.savetxt(log_file_c, cour_list1, cour_list2, delimiter=',', fmt='%s')
        with open(os.path.join(output_dir, 'results.txt'), 'w') as f:
            f.writelines(msgs_all)
        print('result images and text saved to', output_dir)
        print('processing time: %d ms per slice' % int(1000.*total_time/len(slices_to_process)))


def import_tag_data():
    cellname = lambda row, col: '%s%d' % (chr(ord('A') + col - 1), row)
    fn = os.path.join(cfg.PROGDAT_DIR, '%s_%s.xlsx' % ('test_handlabeled', cfg.EXP_NAME))

    wb = load_workbook(fn)
    sheet = wb.active
    tags = []
    thresolds = []
    for p in range(2, sheet.max_row):
        tags.append(sheet[cellname(p, 1)].value)
        thresolds.append(float(sheet[cellname(p, 8)].value))
    assert tags == cfg.runtime_info.tag_list
    cfg.runtime_info.tag_sel_val = torch.tensor(thresolds).to(torch.float)


def load_preprocess_nifti(data):
    vol = (data.get_data().astype('int32') + 32768).astype('uint16')  # to be consistent with png files
    # spacing = -data.get_affine()[0,1]
    # slice_intv = -data.get_affine()[2,2]
    aff = data.get_affine()[:3, :3]
    spacing = np.abs(aff[:2, :2]).max()
    slice_intv = np.abs(aff[2, 2])

    # TODO: Ad-hoc code for normalizing the orientation of the volume.
    # The aim is to make vol[:,:,i] an supine right-left slice
    # It works for the authors' data, but maybe not suitable for some kinds of nifti files
    if np.abs(aff[0, 0]) > np.abs(aff[0, 1]):
        vol = np.transpose(vol, (1, 0, 2))
        aff = aff[[1, 0, 2], :]
    if np.max(aff[0, :2]) > 0:
        vol = vol[::-1, :, :]
    if np.max(aff[1, :2]) > 0:
        vol = vol[:, ::-1, :]
    return vol, spacing, slice_intv


def get_ims(slice_idx, vol, spacing, slice_intv, mask_list):
    num_slice = cfg.INPUT.NUM_SLICES * cfg.INPUT.NUM_IMAGES_3DCE
    im_np, im_scale, crop, mask_list = load_prep_img(vol, slice_idx, spacing, slice_intv, mask_list,
                                          cfg.INPUT.IMG_DO_CLIP, num_slice=num_slice)
    im = im_np - cfg.INPUT.PIXEL_MEAN
    im = torch.from_numpy(im.transpose((2, 0, 1))).to(dtype=torch.float)
    ims = im.split(cfg.INPUT.NUM_IMAGES_3DCE)
    return ims, im_np[:, :, int(num_slice/2)+1], im_scale, crop, mask_list


def gen_output(im, result, info, win_show,  cour_list1, cour_list2, recist_list1, recist_list2):
    im = windowing_rev(im, cfg.INPUT.WINDOWING)
    im = windowing(im, win_show).astype('uint8')
    im = cv2.cvtColor(im, cv2.COLOR_GRAY2RGB)

    scale = cfg.TEST.VISUALIZE.SHOW_SCALE
    im = cv2.resize(im, None, None, fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)

    pred = result.bbox.cpu().numpy()
    labels = result.get_field('labels').cpu().numpy()
    scores = result.get_field('scores').cpu().numpy()
    tag_scores = result.get_field('tag_scores').cpu().numpy()
    tag_predictions = result.get_field('tag_predictions').cpu().numpy()

    mm2pix = info['im_scale'] / info['spacing'] * scale
    contours = result.get_field('contour_mm').cpu().numpy() * mm2pix
    contours = [c[c[:, 0] > 0, :] for c in contours]
    contours = [c+1*scale for c in contours]  # there seems to be a small offset in the mask?
    recists = result.get_field('recist_mm').cpu().numpy() * mm2pix
    recists += 1*scale   # there seems to be a small offset in the mask?
    diameters = result.get_field('diameter_mm').cpu().numpy()

    pred *= scale
    #contour_list = contour_list
    #recist_list = recist_list
    overlay, msgs = draw_results(im, pred,   cour_list1, cour_list2, recist_list1, recist_list2, labels, scores, tag_predictions=tag_predictions, tag_scores=tag_scores,
                                 contours=contours, recists=recists, diameters=diameters)
    overlay = print_msg_on_img(overlay, msgs)
    return overlay, msgs


def print_msg_on_img(overlay, msgs):
    txt_height = 20
    msg_im = np.zeros((txt_height*cfg.TEST.VISUALIZE.DETECTIONS_PER_IMG+10, overlay.shape[1], 3), dtype=np.uint8)
    for p in range(len(msgs)):
        msg = msgs[p].split(' | ')
        msg = msg[0][7:10] + msg[1][:-2] + ': ' + msg[2]
        cv2.putText(msg_im, msg, (0, txt_height*(p+1)),
                    cv2.FONT_HERSHEY_DUPLEX, fontScale=.5,
                    color=(255,255,255), thickness=1)
    return np.vstack((overlay, msg_im))